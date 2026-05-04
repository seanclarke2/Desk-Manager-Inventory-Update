import argparse
import re
import traceback
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    PdfReader = None

try:
    import pytesseract
    from PIL import Image
except Exception:  # pragma: no cover
    pytesseract = None
    Image = None

from playwright.sync_api import sync_playwright


DM_COLUMNS = [
    "Unit",
    "VIN",
    "Vehicle Type",
    "New / Used",
    "Condition",
    "Body Style",
    "Year",
    "Make",
    "Model",
    "Length",
    "Axles",
    "State Registered",
    "Key Number",
    "Title-In",
    "Bill Of Sale Date",
    "Bill of Sale Number",
    "City",
    "Description",
    "ROS / Title Number",
    "Title Status",
    "State",
    "Previous Title Owner",
    "Note",
    "Purchase Cost",
    "Purchase Date",
    "Purchased From",
    "Purchase Method",
    "Purchase Channel",
    "Payment Method",
    "Due Date",
    "Reference No.",
    "Invoice No.",
    "Buyer",
    "Date In",
    "Location",
    "Sold To",
    "Sold Date",
    "Sold Price",
    "Attachment File Name",
]

REPORT_COLUMNS = [
    "Unit",
    "VIN",
    "Import Status",
    "DeskManager Verified",
    "Bill of Sale Uploaded",
    "Uploaded File Name",
    "Issues Found",
    "Needs Manual Review",
    "DeskManager URL if available",
]

SUPPORTED_DOC_EXT = {".pdf", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".webp", ".bmp"}

# Files whose names match this pattern should never be used as valid bill-of-sale sources
# (cancelled, error copies, credit memos, wire instructions, etc.)
BOS_SKIP_PATTERN = re.compile(
    r'cancel|cancelled|canceled|credit|wrong\s*vin|wrong\s*title|unit\s*error|wire\s*instruct',
    re.IGNORECASE,
)

# Bill of Sale folder search order: Sean's (primary/more accurate) → Michelle's (fallback for extras).
# Pass --bos-dir to override. The pipeline will also search the secondary folder automatically.
BOS_DIR_PRIMARY   = Path("/Users/seanclarke/Library/CloudStorage/OneDrive-Personal/sean's  docs/Bill of Sale")
BOS_DIR_SECONDARY = Path("/Users/seanclarke/Library/CloudStorage/OneDrive-Personal/Michelle/PC_Riverside Equipment Sales/Bill of Sales")


def collect_bos_files(bos_dir: Optional[Path] = None) -> List[Path]:
    """Return all valid BOS document paths.

    Searches *bos_dir* (or BOS_DIR_PRIMARY) first, then BOS_DIR_SECONDARY for any
    invoice numbers not already represented, skipping cancelled/error files.
    Files are deduplicated by filename (primary folder wins on collision).
    """
    primary = bos_dir if bos_dir is not None else BOS_DIR_PRIMARY

    def _valid_files(folder: Path) -> List[Path]:
        if not folder.exists():
            return []
        return [
            p for p in sorted(folder.iterdir())
            if p.is_file()
            and p.suffix.lower() in SUPPORTED_DOC_EXT
            and not BOS_SKIP_PATTERN.search(p.name)
        ]

    primary_files = _valid_files(primary)
    primary_names = {f.name for f in primary_files}

    secondary_extras = [
        f for f in _valid_files(BOS_DIR_SECONDARY)
        if f.name not in primary_names
    ]

    return primary_files + secondary_extras

MAKE_ALIASES = {
    "HYUNDAI TRANSLEAD": "Hyundai",
    "HYUNDAI": "Hyundai",
    "UTIL": "Utility",
    "UTILITY": "Utility",
    "UTILITY TRAILER": "Utility",
    "WABASH NATIONAL": "Wabash",
    "WABASH": "Wabash",
    "FREIGHTLINER": "Freightliner",
    "FRT": "Freightliner",
    "GREAT DANE": "Great Dane",
    "STOUGHTON": "Stoughton",
    "CIMC": "CIMC",
}


@dataclass
class ParsedUnit:
    source_file: str
    unit: str
    vin: str
    vehicle_type: str
    new_used: str
    condition: str
    body_style: str
    year: str
    make: str
    model: str
    length: str
    axles: str
    state_registered: str
    key_number: str
    bos_date: str
    bos_number: str
    city: str
    description: str
    ros_title_number: str
    title_status: str
    state: str
    previous_title_owner: str
    note: str
    purchase_cost: str
    purchase_date: str
    purchased_from: str
    purchase_method: str
    purchase_channel: str
    payment_method: str
    due_date: str
    reference_no: str
    invoice_no: str
    buyer: str
    date_in: str
    location: str
    sold_to: str
    sold_date: str


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def normalize_key(text: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean_text(text).lower())


def normalize_vin(vin: str) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", clean_text(vin).upper())


def valid_vin(vin: str) -> bool:
    v = normalize_vin(vin)
    if len(v) != 17:
        return False
    if re.search(r"[IOQ]", v):
        return False
    return bool(re.fullmatch(r"[A-HJ-NPR-Z0-9]{17}", v))


def normalize_unit_for_match(unit: str) -> str:
    s = clean_text(unit).upper()
    if not s:
        return ""
    # Strip G only when it wraps the whole token around digits.
    if re.fullmatch(r"G\d+G", s):
        s = s[1:-1]
    return re.sub(r"\s+", "", s)


def unit_numeric_core(unit: str) -> str:
    s = normalize_unit_for_match(unit)
    m = re.fullmatch(r"[A-Z]*([0-9]+)[A-Z]*", s)
    if not m:
        return ""
    return m.group(1).lstrip("0") or "0"


def parse_date_any(value: str) -> Optional[datetime]:
    v = clean_text(value)
    if not v:
        return None
    patterns = [
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y-%m-%d",
        "%m-%d-%Y",
        "%m-%d-%y",
        "%Y/%m/%d",
        "%b %d, %Y",
        "%B %d, %Y",
    ]
    for p in patterns:
        try:
            return datetime.strptime(v, p)
        except Exception:
            pass
    try:
        return pd.to_datetime(v, errors="coerce").to_pydatetime()
    except Exception:
        return None


def fmt_mmddyyyy(value: str) -> str:
    dt = parse_date_any(value)
    return dt.strftime("%m/%d/%Y") if dt else ""


def fmt_mmddyy_dots(value: str) -> str:
    dt = parse_date_any(value)
    return dt.strftime("%m.%d.%y") if dt else ""


def extract_text_from_pdf(path: Path) -> str:
    if PdfReader is None:
        return ""
    try:
        reader = PdfReader(str(path))
        chunks: List[str] = []
        for page in reader.pages:
            chunks.append(page.extract_text() or "")
        return "\n".join(chunks)
    except Exception:
        return ""


def extract_text_from_image(path: Path) -> str:
    if pytesseract is None or Image is None:
        return ""
    try:
        img = Image.open(path)
        return pytesseract.image_to_string(img) or ""
    except Exception:
        return ""


def extract_doc_text(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return extract_text_from_pdf(path)
    if ext in SUPPORTED_DOC_EXT:
        return extract_text_from_image(path)
    return ""


def detect_body_style(text: str) -> str:
    t = text.lower()
    if "reefer" in t or "refrigerated" in t:
        return "Reefer"
    if "dry van" in t or ("van" in t and "reefer" not in t):
        return "Dry Van"
    if "flatbed" in t or "flat bed" in t:
        return "Flatbed"
    if "chassis" in t:
        return "Chassis"
    if "container" in t:
        return "Container"
    if "dolly" in t:
        return "Dolly"
    if "truck" in t or "tractor" in t:
        return "Truck"
    return "Other"


def detect_make(text: str) -> str:
    up = re.sub(r"\s+", " ", text.upper())
    for k, v in MAKE_ALIASES.items():
        if k in up:
            return v
    m = re.search(r"\b(MAKE|MFR|MANUFACTURER)\s*[:#-]?\s*([A-Z0-9][A-Z0-9\- ]{1,30})", up)
    if m:
        candidate = m.group(2).strip()
        return MAKE_ALIASES.get(candidate, candidate.title())
    return ""


def detect_length(text: str) -> str:
    t = text.upper()
    m_combo = re.search(r"\b(20\s*/\s*40|40\s*/\s*45)\b", t)
    if m_combo:
        return re.sub(r"\s+", "", m_combo.group(1))
    m = re.search(r"\b(20|24|28|32|40|45|48|53)\s*(?:'|FT|FEET)\b", t)
    if m:
        return m.group(1)
    return ""


def detect_axles(text: str) -> str:
    t = text.lower()
    if "tandem" in t:
        return "Tandem"
    m = re.search(r"\b([345678])\s*axle\b", t)
    if m:
        return m.group(1)
    return ""


def detect_city(raw_city: str) -> Tuple[str, str]:
    city = clean_text(raw_city)
    note = ""
    if not city:
        return "", "City missing"
    if city.upper() == "CRST CA":
        return "Jurupa Valley", ""
    if re.fullmatch(r"[A-Z]{2,6}(\s+[A-Z]{2,6})?", city.upper()):
        note = f"City uses abbreviation '{city}', verify"
    return city, note


def detect_doc_header_fields(text: str, filename: str) -> Dict[str, str]:
    compact = re.sub(r"\s+", " ", text)

    date_patterns = [
        r"(?:bill\s*of\s*sale\s*date|date)\s*[:#-]?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
        r"(\d{4}-\d{2}-\d{2})",
    ]
    bos_date = ""
    for p in date_patterns:
        m = re.search(p, compact, flags=re.IGNORECASE)
        if m:
            bos_date = fmt_mmddyyyy(m.group(1))
            if bos_date:
                break

    bos_num = ""
    num_patterns = [
        r"(?:bill\s*of\s*sale\s*(?:#|number|no\.?))\s*[:#-]?\s*([A-Za-z0-9\-/]+)",
        r"(?:invoice\s*(?:#|number|no\.?))\s*[:#-]?\s*([A-Za-z0-9\-/]+)",
        r"(?:document\s*(?:#|number|no\.?))\s*[:#-]?\s*([A-Za-z0-9\-/]+)",
    ]
    for p in num_patterns:
        m = re.search(p, compact, flags=re.IGNORECASE)
        if m:
            bos_num = clean_text(m.group(1))
            break

    if not bos_num:
        fm = re.search(r"(?:^|[_\s-])(\d{5,8})(?:[_\s.-]|$)", filename)
        if fm:
            bos_num = fm.group(1)

    city = ""
    m_city = re.search(r"\b(?:city|location)\s*[:#-]?\s*([A-Za-z\s]{2,40})\b", compact, flags=re.IGNORECASE)
    if m_city:
        city = clean_text(m_city.group(1))

    seller = ""
    m_seller = re.search(r"\b(?:seller|vendor|purchased\s*from)\s*[:#-]?\s*([A-Za-z0-9&.,'\-\s]{2,80})", compact, flags=re.IGNORECASE)
    if m_seller:
        seller = clean_text(m_seller.group(1))

    amount = ""
    m_amt = re.search(r"(?:total|purchase\s*price|amount)\s*[:#-]?\s*\$?\s*([0-9][0-9,]*(?:\.\d{1,2})?)", compact, flags=re.IGNORECASE)
    if m_amt:
        amount = m_amt.group(1).replace(",", "")

    return {
        "bos_date": bos_date,
        "bos_number": bos_num,
        "city": city,
        "seller": seller,
        "purchase_cost": amount,
    }


def extract_units_from_text(text: str, source_file: str) -> List[Dict[str, str]]:
    compact = re.sub(r"\s+", " ", text)
    header = detect_doc_header_fields(compact, source_file)

    vins = []
    for m in re.finditer(r"\b[A-HJ-NPR-Z0-9]{17}\b", compact.upper()):
        vin = m.group(0)
        if valid_vin(vin):
            vins.append((vin, m.start()))

    if not vins:
        # Keep one fallback row so this document appears in Needs Review.
        vins = [("", 0)]

    results: List[Dict[str, str]] = []
    for vin, pos in vins:
        window = compact[max(0, pos - 250): pos + 350]

        m_unit = re.search(r"\b(?:unit|stock)\s*(?:#|number|no\.)?\s*[:#-]?\s*([A-Za-z0-9\-\/]{2,20})", window, flags=re.IGNORECASE)
        unit = clean_text(m_unit.group(1) if m_unit else "")
        if not unit and source_file:
            fn_unit = re.search(r"\b(?:unit|stock)[_\s-]*([A-Za-z0-9\-]{2,20})\b", source_file, flags=re.IGNORECASE)
            if fn_unit:
                unit = clean_text(fn_unit.group(1))

        year = ""
        m_year = re.search(r"\b(19[9][0-9]|20[0-3][0-9])\b", window)
        if m_year:
            year = m_year.group(1)

        make = detect_make(window)
        body_style = detect_body_style(window)
        length = detect_length(window)
        axles = detect_axles(window)

        model = ""
        if body_style == "Dry Van" and length:
            model = f"{length}' Tandem Dry Van" if axles == "Tandem" else f"{length}' Dry Van"
        elif body_style == "Reefer" and length:
            model = f"{length}' Tandem Reefer" if axles == "Tandem" else f"{length}' Reefer"
        elif body_style == "Chassis" and length:
            if "/" in length:
                model = f"{length} Combo Chassis"
            else:
                model = f"{length}' Chassis"

        city, city_note = detect_city(header["city"])
        note_bits = [city_note] if city_note else []

        bos_date = header["bos_date"]
        key_number = fmt_mmddyy_dots(bos_date)
        if not key_number:
            note_bits.append("Missing Bill Of Sale Date; Key Number left blank")

        purchase_cost = header["purchase_cost"]
        if not purchase_cost:
            purchase_cost = "1"
            note_bits.append("Price missing, entered as $1")

        description_lines = []
        short_model = model or body_style or "Trailer"
        if year or make or short_model:
            description_lines.append(" ".join([p for p in [year, make, short_model] if p]))
        if vin:
            description_lines.append(f"VIN: {vin}")
        if bos_date:
            description_lines.append(f"Bill of Sale Date: {bos_date}")
        if header["bos_number"]:
            description_lines.append(f"Bill of Sale #: {header['bos_number']}")

        results.append(
            {
                "Unit": unit,
                "VIN": vin,
                "Vehicle Type": "Trailer",
                "New / Used": "Used",
                "Condition": "Used",
                "Body Style": body_style,
                "Year": year,
                "Make": make,
                "Model": model,
                "Length": length,
                "Axles": axles,
                "State Registered": "",
                "Key Number": key_number,
                "Title-In": "",
                "Bill Of Sale Date": bos_date,
                "Bill of Sale Number": header["bos_number"],
                "City": city,
                "Description": "\n".join(description_lines),
                "ROS / Title Number": "",
                "Title Status": "",
                "State": "",
                "Previous Title Owner": "",
                "Note": "; ".join([n for n in note_bits if n]),
                "Purchase Cost": purchase_cost,
                "Purchase Date": bos_date,
                "Purchased From": header["seller"],
                "Purchase Method": header["seller"],
                "Purchase Channel": "",
                "Payment Method": "",
                "Due Date": "",
                "Reference No.": header["bos_number"],
                "Invoice No.": "",
                "Buyer": "",
                "Date In": bos_date,
                "Location": "",
                "Sold To": "",
                "Sold Date": "",
                "Attachment File Name": source_file,
            }
        )

    return results


def load_database_excel(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    frames: List[pd.DataFrame] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [clean_text(v) for v in rows[0]]
        if not any(headers):
            continue
        data = rows[1:]
        df = pd.DataFrame(data, columns=headers)
        if len(df.columns) == 0:
            continue
        frames.append(df)
    if not frames:
        raise ValueError("No readable data found in database Excel workbook.")
    db = pd.concat(frames, ignore_index=True)
    db.columns = [clean_text(c) for c in db.columns]
    return db.fillna("")


def find_db_col(df: pd.DataFrame, choices: List[str]) -> Optional[str]:
    by_key = {normalize_key(c): c for c in df.columns}
    for choice in choices:
        if normalize_key(choice) in by_key:
            return by_key[normalize_key(choice)]
    for c in df.columns:
        nk = normalize_key(c)
        if any(normalize_key(choice) in nk for choice in choices):
            return c
    return None


def build_db_indexes(db: pd.DataFrame) -> Tuple[Dict[str, List[int]], Dict[str, List[int]], Dict[str, List[int]], Dict[str, str]]:
    vin_col = find_db_col(db, ["VIN", "Vehicle Identification Number"])
    unit_col = find_db_col(db, ["Unit", "Stock", "Stock Number", "Unit Number"])

    if not vin_col and not unit_col:
        raise ValueError("Could not detect VIN or Unit column in database workbook.")

    idx_vin: Dict[str, List[int]] = defaultdict(list)
    idx_unit: Dict[str, List[int]] = defaultdict(list)
    idx_core: Dict[str, List[int]] = defaultdict(list)

    for i, row in db.iterrows():
        vin = normalize_vin(row.get(vin_col, "") if vin_col else "")
        unit = clean_text(row.get(unit_col, "") if unit_col else "")
        if vin:
            idx_vin[vin].append(i)
        if unit:
            nu = normalize_unit_for_match(unit)
            if nu:
                idx_unit[nu].append(i)
            core = unit_numeric_core(unit)
            if core:
                idx_core[core].append(i)

    meta = {
        "vin_col": vin_col or "",
        "unit_col": unit_col or "",
        "year_col": find_db_col(db, ["Year"]),
        "make_col": find_db_col(db, ["Make", "Manufacturer"]),
        "model_col": find_db_col(db, ["Model"]),
        "body_col": find_db_col(db, ["Body Style", "Body"]),
        "length_col": find_db_col(db, ["Length"]),
        "axles_col": find_db_col(db, ["Axles", "Axels", "Axle"]),
        "state_reg_col": find_db_col(db, ["State Registered", "State"]),
        "purchase_cost_col": find_db_col(db, ["Purchase Cost", "Cost", "Price"]),
        "purchase_date_col": find_db_col(db, ["Purchase Date", "Bill Of Sale Date"]),
        "purchased_from_col": find_db_col(db, ["Purchased From", "Vendor", "Seller"]),
    }
    return idx_vin, idx_unit, idx_core, meta


def choose_single(candidates: List[int]) -> Tuple[Optional[int], str]:
    if not candidates:
        return None, ""
    unique = sorted(set(candidates))
    if len(unique) == 1:
        return unique[0], ""
    return None, "Multiple database matches"


def apply_db_verification(row: Dict[str, str], db: pd.DataFrame, idx_vin, idx_unit, idx_core, meta: Dict[str, str]) -> Tuple[Dict[str, str], bool]:
    note_parts = [p for p in clean_text(row.get("Note", "")).split(";") if p.strip()]
    needs_review = False

    vin = normalize_vin(row.get("VIN", ""))
    unit = clean_text(row.get("Unit", ""))
    unit_norm = normalize_unit_for_match(unit)

    match_idx = None
    match_reason = ""

    if vin:
        match_idx, match_reason = choose_single(idx_vin.get(vin, []))
    if match_idx is None and unit_norm:
        match_idx, match_reason = choose_single(idx_unit.get(unit_norm, []))
    if match_idx is None and unit_norm:
        match_idx, match_reason = choose_single(idx_core.get(unit_numeric_core(unit_norm), []))

    if match_reason:
        needs_review = True
        note_parts.append(match_reason)

    db_row = db.loc[match_idx] if match_idx is not None else None

    # If unit is still missing and chassis with VIN, use C-last4 rule.
    if not unit and row.get("Body Style", "") == "Chassis" and vin and len(vin) >= 4:
        row["Unit"] = f"C-{vin[-4:]}"

    def fill_from_db(target: str, db_col_key: str):
        nonlocal needs_review
        if db_row is None:
            return
        db_col = meta.get(db_col_key, "")
        if not db_col:
            return
        db_val = clean_text(db_row.get(db_col, ""))
        if not db_val:
            return
        cur_val = clean_text(row.get(target, ""))
        if not cur_val:
            row[target] = db_val
            return
        if target == "Make":
            if clean_text(cur_val).lower() != clean_text(db_val).lower():
                note_parts.append(f"Bill of Sale vs DB conflict for {target}: BOS='{cur_val}' DB='{db_val}'")
                needs_review = True
        elif target in {"Model", "Body Style", "Length", "Axles", "Year"}:
            if normalize_key(cur_val) != normalize_key(db_val):
                note_parts.append(f"Bill of Sale vs DB conflict for {target}: BOS='{cur_val}' DB='{db_val}'")
                needs_review = True

    fill_from_db("Unit", "unit_col")
    fill_from_db("Year", "year_col")
    fill_from_db("Make", "make_col")
    fill_from_db("Model", "model_col")
    fill_from_db("Body Style", "body_col")
    fill_from_db("Length", "length_col")
    fill_from_db("Axles", "axles_col")
    fill_from_db("State Registered", "state_reg_col")
    fill_from_db("Purchase Cost", "purchase_cost_col")
    fill_from_db("Purchase Date", "purchase_date_col")
    fill_from_db("Purchased From", "purchased_from_col")

    # Keep Purchase Method aligned to source/seller if currently blank.
    if not clean_text(row.get("Purchase Method", "")) and clean_text(row.get("Purchased From", "")):
        row["Purchase Method"] = clean_text(row["Purchased From"])

    # VIN validation after DB verify.
    if not valid_vin(row.get("VIN", "")):
        needs_review = True
        note_parts.append("VIN missing/invalid after DB verification")

    if not clean_text(row.get("Unit", "")):
        needs_review = True
        note_parts.append("Unit missing after Bill of Sale + DB check")

    if not clean_text(row.get("Bill Of Sale Date", "")):
        needs_review = True
        note_parts.append("Bill Of Sale Date missing")

    if not clean_text(row.get("Key Number", "")):
        needs_review = True
        note_parts.append("Key Number missing because BOS date unavailable")

    if not clean_text(row.get("City", "")):
        needs_review = True
        note_parts.append("City missing")

    row["Note"] = "; ".join(dict.fromkeys([p.strip() for p in note_parts if p.strip()]))
    return row, needs_review


def build_import_rows(bos_dir: Path, db_xlsx: Path, log_lines: List[str]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if not db_xlsx.exists():
        raise FileNotFoundError(f"Database Excel file not found: {db_xlsx}")

    db = load_database_excel(db_xlsx)
    idx_vin, idx_unit, idx_core, meta = build_db_indexes(db)

    docs = collect_bos_files(bos_dir)
    if not docs:
        raise ValueError("No valid Bill of Sale PDFs/images found (checked primary and secondary folders).")

    rows: List[Dict[str, str]] = []
    review_rows: List[Dict[str, str]] = []

    for path in docs:
        text = extract_doc_text(path)
        if not clean_text(text):
            log_lines.append(f"WARN: Could not extract text from {path.name}; adding review placeholder.")
            parsed = extract_units_from_text("", path.name)
        else:
            parsed = extract_units_from_text(text, path.name)

        for r in parsed:
            # Enforce exact output columns and defaults.
            row = {c: clean_text(r.get(c, "")) for c in DM_COLUMNS}
            row["Vehicle Type"] = row["Vehicle Type"] or "Trailer"
            row["New / Used"] = row["New / Used"] or "Used"
            row["Condition"] = row["Condition"] or "Used"
            row["Bill Of Sale Date"] = fmt_mmddyyyy(row["Bill Of Sale Date"])
            row["Purchase Date"] = fmt_mmddyyyy(row["Purchase Date"]) if row.get("Purchase Date") else row.get("Bill Of Sale Date", "")
            row["Date In"] = fmt_mmddyyyy(row["Date In"]) if row.get("Date In") else row.get("Bill Of Sale Date", "")
            row["Key Number"] = fmt_mmddyy_dots(row["Bill Of Sale Date"]) if row.get("Bill Of Sale Date") else ""

            row, needs_review = apply_db_verification(row, db, idx_vin, idx_unit, idx_core, meta)

            if not clean_text(row.get("Purchase Cost", "")):
                row["Purchase Cost"] = "1"
                row["Note"] = (row["Note"] + "; Price missing, entered as $1").strip("; ")

            rows.append(row)
            if needs_review:
                review_rows.append(row.copy())

    import_df = pd.DataFrame(rows)
    needs_review_df = pd.DataFrame(review_rows)

    for col in DM_COLUMNS:
        if col not in import_df.columns:
            import_df[col] = ""
        if col not in needs_review_df.columns:
            needs_review_df[col] = ""

    import_df = import_df[DM_COLUMNS]
    needs_review_df = needs_review_df[DM_COLUMNS]

    return import_df, needs_review_df


def dataframe_to_vehicle_fill_csv(df: pd.DataFrame, path: Path):
    mapped = df.copy()
    mapped["Stock Number"] = mapped["Unit"]
    mapped["Bill Of Sale Number"] = mapped["Bill of Sale Number"]
    mapped.to_csv(path, index=False)


def run_import_tool_upload(page, import_xlsx: Path):
    candidate_urls = [
        "https://dm.automanager.com/InventoryImport",
        "https://dm.automanager.com/Inventory/Import",
        "https://dm.automanager.com/Inventory?action=import",
    ]

    landed = False
    for url in candidate_urls:
        try:
            page.goto(url, wait_until="domcontentloaded")
            page.wait_for_timeout(2000)
            if page.locator("input[type='file']").count() > 0:
                landed = True
                break
        except Exception:
            continue

    if not landed:
        raise Exception("Could not locate DeskManager inventory import page/file input.")

    file_input = page.locator("input[type='file']").first
    file_input.set_input_files(str(import_xlsx))
    page.wait_for_timeout(1500)

    for text in ["Upload", "Next", "Continue"]:
        btn = page.locator(f"button:has-text('{text}'), input[value*='{text}' i]").first
        if btn.count() > 0 and btn.is_visible() and btn.is_enabled():
            btn.click()
            page.wait_for_timeout(1500)

    mapping_page_text = ""
    try:
        mapping_page_text = page.inner_text("body")
    except Exception:
        pass

    present = 0
    for col in DM_COLUMNS:
        if col.lower() in mapping_page_text.lower():
            present += 1

    if present < int(len(DM_COLUMNS) * 0.6):
        raise Exception(
            "Import mapping validation failed. Expected columns are not visible in mapping step; stopping for safety."
        )

    for text in ["Start Import", "Import", "Finish", "Done"]:
        btn = page.locator(f"button:has-text('{text}'), input[value*='{text}' i]").first
        if btn.count() > 0 and btn.is_visible() and btn.is_enabled():
            btn.click()
            page.wait_for_timeout(1500)


def import_and_attach_in_deskmanager(
    import_df: pd.DataFrame,
    bos_dir: Optional[Path],
    log_lines: List[str],
    import_xlsx: Path,
) -> pd.DataFrame:
    import deskmanager_vehicle_fill as dvf
    import deskmanager_bill_of_sales_upload as bos_upload

    results: List[Dict[str, str]] = []

    with sync_playwright() as p:
        browser, context, page = dvf.create_authenticated_session(p)
        try:
            run_import_tool_upload(page, import_xlsx)
            log_lines.append(f"INFO: Import spreadsheet uploaded: {import_xlsx}")

            for _, row in import_df.iterrows():
                unit = clean_text(row.get("Unit", ""))
                vin = clean_text(row.get("VIN", ""))
                attachment_name = clean_text(row.get("Attachment File Name", ""))
                # Find attachment in primary then secondary folder (skipping cancelled/error files)
                attachment_path: Optional[Path] = None
                if attachment_name:
                    for folder in [bos_dir or BOS_DIR_PRIMARY, BOS_DIR_SECONDARY]:
                        candidate = folder / attachment_name
                        if candidate.is_file() and not BOS_SKIP_PATTERN.search(candidate.name):
                            attachment_path = candidate
                            break

                status = "Failed"
                verified = "No"
                uploaded = "No"
                issue = ""
                dm_url = ""

                try:
                    if not unit:
                        status = "Needs Review"
                        issue = "Missing Unit"
                    else:
                        dvf.open_vehicle_form(page, unit)
                        status = "Imported"
                        verified = "Yes"
                        dm_url = page.url

                        if attachment_path and attachment_path.exists():
                            up_ok = bos_upload.upload_bill_of_sale_to_attachments(page, attachment_path)
                            uploaded = "Yes" if up_ok else "No"
                            if not up_ok:
                                issue = "Attachment upload failed"
                        else:
                            issue = "Attachment file not found"
                            uploaded = "No"

                except Exception as exc:
                    issue = str(exc)
                    log_lines.append(f"ERROR unit={unit or '-'} vin={vin or '-'}: {exc}")

                needs_manual_review = "Yes" if status in {"Failed", "Needs Review"} or uploaded == "No" else "No"
                results.append(
                    {
                        "Unit": unit,
                        "VIN": vin,
                        "Import Status": status,
                        "DeskManager Verified": verified,
                        "Bill of Sale Uploaded": uploaded,
                        "Uploaded File Name": attachment_name,
                        "Issues Found": issue,
                        "Needs Manual Review": needs_manual_review,
                        "DeskManager URL if available": dm_url,
                    }
                )
        finally:
            try:
                browser.close()
            except Exception:
                pass

    report_df = pd.DataFrame(results)
    for col in REPORT_COLUMNS:
        if col not in report_df.columns:
            report_df[col] = ""
    return report_df[REPORT_COLUMNS]


def build_not_executed_report(import_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in import_df.iterrows():
        rows.append(
            {
                "Unit": clean_text(row.get("Unit", "")),
                "VIN": clean_text(row.get("VIN", "")),
                "Import Status": "Needs Review",
                "DeskManager Verified": "No",
                "Bill of Sale Uploaded": "No",
                "Uploaded File Name": clean_text(row.get("Attachment File Name", "")),
                "Issues Found": "DeskManager execution not run (set --run-deskmanager to execute import/update + attachment upload)",
                "Needs Manual Review": "Yes",
                "DeskManager URL if available": "",
            }
        )
    return pd.DataFrame(rows, columns=REPORT_COLUMNS)


def write_outputs(output_dir: Path, import_df: pd.DataFrame, review_df: pd.DataFrame, report_df: pd.DataFrame, log_lines: List[str]):
    output_dir.mkdir(parents=True, exist_ok=True)

    import_xlsx = output_dir / "DeskManager_Import_Ready.xlsx"
    review_xlsx = output_dir / "Needs_Review.xlsx"
    report_xlsx = output_dir / "DeskManager_Verification_Report.xlsx"
    log_txt = output_dir / "Import_Log.txt"

    import_df.to_excel(import_xlsx, index=False)
    review_df.to_excel(review_xlsx, index=False)
    report_df.to_excel(report_xlsx, index=False)

    with open(log_txt, "w", encoding="utf-8") as f:
        if not log_lines:
            f.write("No warnings or errors logged.\n")
        else:
            f.write("\n".join(log_lines) + "\n")



def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="DeskManager Bill of Sale import + attachment automation pipeline")
    parser.add_argument("--bos-dir", default=None, help="Folder containing Bill of Sale PDFs/images (default: Sean's OneDrive folder, with Michelle's as fallback)")
    parser.add_argument("--database-xlsx", required=True, help="Database Excel file path")
    parser.add_argument("--output-dir", default=str(Path.cwd()), help="Output folder for generated files")
    parser.add_argument("--run-deskmanager", action="store_true", help="Execute DeskManager update/import + attachment uploads")
    return parser.parse_args()


def main():
    args = parse_args()
    bos_dir = Path(args.bos_dir).expanduser().resolve() if args.bos_dir else None
    db_xlsx = Path(args.database_xlsx).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()

    log_lines: List[str] = []

    try:
        import_df, review_df = build_import_rows(bos_dir, db_xlsx, log_lines)

        output_dir.mkdir(parents=True, exist_ok=True)
        import_xlsx = output_dir / "DeskManager_Import_Ready.xlsx"
        review_xlsx = output_dir / "Needs_Review.xlsx"
        import_df.to_excel(import_xlsx, index=False)
        review_df.to_excel(review_xlsx, index=False)

        if args.run_deskmanager:
            report_df = import_and_attach_in_deskmanager(import_df, bos_dir, log_lines, import_xlsx)
        else:
            report_df = build_not_executed_report(import_df)

        write_outputs(output_dir, import_df, review_df, report_df, log_lines)

        print("Pipeline complete.")
        print(f"Generated: {output_dir / 'DeskManager_Import_Ready.xlsx'}")
        print(f"Generated: {output_dir / 'Needs_Review.xlsx'}")
        print(f"Generated: {output_dir / 'DeskManager_Verification_Report.xlsx'}")
        print(f"Generated: {output_dir / 'Import_Log.txt'}")

    except Exception as exc:
        print(f"ERROR: {exc}")
        traceback.print_exc()
        raise


if __name__ == "__main__":
    main()
