import os
import re
import sys
import traceback
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright

BASE_URL = "https://dm.automanager.com/Account/Login?ReturnUrl=%2fDashboard"
INVENTORY_URL = "https://dm.automanager.com/Inventory?action=10"

# Credentials must come from environment variables.
AUTOMANAGER_USERNAME = os.getenv("AUTOMANAGER_USERNAME", "").strip()
AUTOMANAGER_PASSWORD = os.getenv("AUTOMANAGER_PASSWORD", "").strip()

HEADLESS = os.getenv("PLAYWRIGHT_HEADLESS", "false").lower() in {"1", "true", "yes"}
MAX_UNITS = int(os.getenv("DESKMANAGER_MAX_UNITS", "0"))  # 0 = no limit

DB_FILENAME = "2026 Database_Share_2026-03-25.xlsx"
DEFAULT_OUTPUT = "DeskManager_Matched_Inventory.xlsx"
DEFAULT_ERROR_LOG = "DeskManager_Match_Errors.txt"

DATABASE_XLSX = os.getenv("DESKMANAGER_DATABASE_XLSX", "").strip()
OUTPUT_XLSX = os.getenv("DESKMANAGER_OUTPUT_XLSX", DEFAULT_OUTPUT).strip()
ERROR_LOG_PATH = os.getenv("DESKMANAGER_ERROR_LOG", DEFAULT_ERROR_LOG).strip()

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(line_buffering=True)
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(line_buffering=True)


@dataclass
class DatabaseData:
    headers: List[str]
    rows: List[Dict[str, str]]
    stock_header: str
    vin_header: str


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_text(value).lower())


def normalize_vin(vin: str) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", normalize_text(vin).upper())


def is_valid_vin(vin: str) -> bool:
    v = normalize_vin(vin)
    if len(v) != 17:
        return False
    if re.search(r"[IOQ]", v):
        return False
    return bool(re.fullmatch(r"[A-HJ-NPR-Z0-9]{17}", v))


def normalize_stock_exact(stock: str) -> str:
    return normalize_text(stock).upper()


def stock_numeric_core(stock: str) -> str:
    """
    Extract numeric core only when stock contains optional letters at front/back.
    Does not strip letters in the middle.
    Examples:
      G3009G -> 3009
      2276HL -> 2276
      AB12CD -> 12
      12AB34 -> '' (letters in middle, do not strip)
    """
    s = normalize_text(stock).upper()
    m = re.fullmatch(r"[A-Z]*([0-9]+)[A-Z]*", s)
    if not m:
        return ""
    digits = m.group(1).lstrip("0")
    return digits or "0"


def find_database_file() -> Path:
    if DATABASE_XLSX:
        p = Path(DATABASE_XLSX).expanduser()
        if p.exists():
            return p
        raise FileNotFoundError(f"Database file not found at DESKMANAGER_DATABASE_XLSX: {p}")

    candidates = [
        Path.cwd() / DB_FILENAME,
        Path.home() / "Library/CloudStorage/OneDrive-Personal" / DB_FILENAME,
        Path.home() / "Library/CloudStorage/OneDrive-Personal/Michelle" / DB_FILENAME,
    ]
    for c in candidates:
        if c.exists():
            return c

    onedrive_root = Path.home() / "Library/CloudStorage/OneDrive-Personal"
    if onedrive_root.exists():
        matches = list(onedrive_root.rglob(DB_FILENAME))
        if matches:
            return matches[0]

    raise FileNotFoundError(
        f"Could not locate {DB_FILENAME}. Set DESKMANAGER_DATABASE_XLSX to the full path."
    )


def detect_header(headers: List[str], kind: str) -> str:
    key_map = {h: normalize_key(h) for h in headers}

    if kind == "vin":
        preferred = {
            "vin",
            "vehicleidentificationnumber",
            "vehicleidnumber",
        }
        contains = ["vin", "vehicleidentification"]
    else:
        preferred = {
            "stock",
            "stocknumber",
            "stockno",
            "stock#",
            "unit",
            "unitnumber",
            "unitno",
            "unit#",
        }
        contains = ["stock", "unit"]

    for h, k in key_map.items():
        if k in preferred:
            return h

    for h, k in key_map.items():
        if any(token in k for token in contains):
            return h

    raise ValueError(f"Unable to detect {kind.upper()} column in Excel database headers: {headers}")


def load_excel_database(path: Path) -> DatabaseData:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("Excel database file appears empty.") from exc

    raw_headers = [normalize_text(h) for h in header_row]
    if not any(raw_headers):
        raise ValueError("Excel database header row is empty.")

    # Deduplicate headers: when a header name appears more than once, keep only
    # the FIRST occurrence so that row dicts use the primary column, not a later
    # duplicate section in the spreadsheet.
    seen_hdr: set = set()
    first_occurrence_indices: List[int] = []
    deduped_headers: List[str] = []
    for i, h in enumerate(raw_headers):
        if h not in seen_hdr:
            seen_hdr.add(h)
            first_occurrence_indices.append(i)
            deduped_headers.append(h)

    headers = deduped_headers

    rows: List[Dict[str, str]] = []
    for row in rows_iter:
        values = [normalize_text(v) for v in row]
        if not any(values):
            continue
        record = {
            headers[j]: values[i] if i < len(values) else ""
            for j, i in enumerate(first_occurrence_indices)
        }
        rows.append(record)

    stock_header = detect_header(headers, kind="stock")
    vin_header = detect_header(headers, kind="vin")

    return DatabaseData(
        headers=headers,
        rows=rows,
        stock_header=stock_header,
        vin_header=vin_header,
    )


def build_indexes(db: DatabaseData):
    vin_idx: Dict[str, List[Dict[str, str]]] = {}
    stock_exact_idx: Dict[str, List[Dict[str, str]]] = {}
    stock_numeric_idx: Dict[str, List[Dict[str, str]]] = {}

    for row in db.rows:
        vin = normalize_vin(row.get(db.vin_header, ""))
        if vin:
            vin_idx.setdefault(vin, []).append(row)

        stock = normalize_stock_exact(row.get(db.stock_header, ""))
        if stock:
            stock_exact_idx.setdefault(stock, []).append(row)

        core = stock_numeric_core(row.get(db.stock_header, ""))
        if core:
            stock_numeric_idx.setdefault(core, []).append(row)

    return vin_idx, stock_exact_idx, stock_numeric_idx


def login(page):
    print("Navigating to login page...")
    page.goto(BASE_URL, wait_until="networkidle")
    page.wait_for_selector('input[type="text"], input[type="email"]', timeout=10000)

    username_input = page.locator('input[type="text"], input[type="email"]').first
    password_input = page.locator('input[type="password"]').first

    print("Filling login credentials...")
    username_input.fill(AUTOMANAGER_USERNAME)
    page.wait_for_timeout(400)
    password_input.fill(AUTOMANAGER_PASSWORD)
    page.wait_for_timeout(400)

    login_button = page.locator('button:has-text("Log in"), button:has-text("LOGIN"), input[type="submit"]').first
    login_button.click()
    page.wait_for_url("**/Dashboard**", timeout=30000)
    print("✓ Login successful")


def navigate_to_inventory(page):
    print("Navigating to Inventory...")
    page.goto(INVENTORY_URL, wait_until="networkidle")
    page.wait_for_timeout(2500)
    print("✓ On Inventory page")


def get_vehicle_hrefs_from_inventory(page) -> List[str]:
    seen = set()
    vehicle_hrefs: List[str] = []
    page_num = 1

    while True:
        page.wait_for_timeout(2500)
        all_links = page.locator('a[href]').all()

        before = len(seen)
        for link in all_links:
            try:
                href = (link.get_attribute("href") or "").strip()
                if not href:
                    continue
                href = href.split("#")[0]
                href_lower = href.lower()
                if not any(t in href_lower for t in ["/inventory/details/", "/inventory/edit/"]):
                    continue
                if href in seen:
                    continue
                seen.add(href)
                if not href.startswith("http"):
                    href = f"https://dm.automanager.com{href}"
                vehicle_hrefs.append(href)
            except Exception:
                pass

        added = len(seen) - before
        print(f"  Page {page_num}: +{added} vehicles (total so far: {len(vehicle_hrefs)})")

        next_btn = None
        for next_sel in [
            'li:not([style*="display:none"]):not([style*="display: none"]) > a:text-is("Next")',
            'a:text-is("Next")',
            'li.active + li > a',
            'a[aria-label="Next"]',
        ]:
            try:
                el = page.locator(next_sel).last
                if el.count() > 0 and el.is_visible() and el.is_enabled():
                    next_btn = el
                    break
            except Exception:
                pass

        if next_btn:
            next_btn.click()
            page_num += 1
            page.wait_for_timeout(3000)
        else:
            break

    print(f"Found {len(vehicle_hrefs)} unique vehicle pages across {page_num} page(s)")
    return vehicle_hrefs


def read_label_field(page, label_text: str) -> Optional[str]:
    try:
        label = page.get_by_text(label_text, exact=True).first
        if not label.is_visible():
            return None

        for_id = label.evaluate("el => el.htmlFor || el.getAttribute('for') || ''")
        if for_id:
            inp = page.locator(f"#{for_id}")
            if inp.count() > 0:
                v = inp.first.input_value().strip()
                return v or None

        val = label.evaluate(
            """
            el => {
                let sib = el.nextElementSibling;
                while (sib) {
                    if (sib.tagName === 'INPUT' || sib.tagName === 'SELECT' || sib.tagName === 'TEXTAREA') return sib.value || sib.textContent || '';
                    let inp = sib.querySelector('input, select, textarea');
                    if (inp) return inp.value || inp.textContent || '';
                    sib = sib.nextElementSibling;
                }
                let row = el.closest('tr, div');
                if (row) {
                    let inp = row.querySelector('input, select, textarea');
                    if (inp) return inp.value || inp.textContent || '';
                }
                return '';
            }
            """
        )
        val = normalize_text(val)
        return val or None
    except Exception:
        return None


def get_stock_vin_from_vehicle(page) -> Dict[str, str]:
    stock = read_label_field(page, "Stock Number")
    vin = read_label_field(page, "VIN")

    if not stock:
        try:
            el = page.locator('input[id*="stock" i], input[name*="stock" i]').first
            if el.count() > 0:
                stock = normalize_text(el.input_value())
        except Exception:
            pass

    if not vin:
        try:
            el = page.locator('input[id*="vin" i], input[name*="vin" i]').first
            if el.count() > 0:
                vin = normalize_text(el.input_value())
        except Exception:
            pass

    return {
        "stock": stock or "",
        "vin": vin or "",
    }


def evaluate_match(
    dm_stock: str,
    dm_vin: str,
    db: DatabaseData,
    vin_idx,
    stock_exact_idx,
    stock_numeric_idx,
):
    dm_stock_exact = normalize_stock_exact(dm_stock)
    dm_vin_norm = normalize_vin(dm_vin)
    dm_vin_valid = is_valid_vin(dm_vin_norm)

    def result(status: str, method: str, notes: str = "", matched_row: Optional[Dict[str, str]] = None):
        return {
            "status": status,
            "method": method,
            "notes": notes,
            "row": matched_row,
        }

    if dm_vin_valid:
        vin_candidates = vin_idx.get(dm_vin_norm, [])
        if len(vin_candidates) > 1:
            return result(
                status="Needs Review - Multiple Possible Matches",
                method="VIN",
                notes=f"Multiple Excel rows share VIN {dm_vin_norm}.",
            )
        if len(vin_candidates) == 1:
            row = vin_candidates[0]
            excel_stock = normalize_stock_exact(row.get(db.stock_header, ""))
            if dm_stock_exact and excel_stock and dm_stock_exact != excel_stock:
                return result(
                    status="Found - VIN Match / Stock Number Different",
                    method="VIN",
                    notes=f"DeskManager stock '{dm_stock}' differs from Excel stock '{row.get(db.stock_header, '')}'.",
                    matched_row=row,
                )
            return result(
                status="Found - VIN Match",
                method="VIN",
                matched_row=row,
            )

    exact_candidates = stock_exact_idx.get(dm_stock_exact, []) if dm_stock_exact else []
    if len(exact_candidates) > 1:
        return result(
            status="Needs Review - Multiple Possible Matches",
            method="Stock Exact",
            notes=f"Multiple Excel rows matched stock '{dm_stock}'.",
        )
    if len(exact_candidates) == 1:
        row = exact_candidates[0]
        excel_vin = normalize_vin(row.get(db.vin_header, ""))
        if dm_vin_valid and excel_vin and dm_vin_norm != excel_vin:
            return result(
                status="Needs Review - Stock Match But VIN Conflict",
                method="Stock Exact",
                notes=f"DeskManager VIN '{dm_vin_norm}' conflicts with Excel VIN '{excel_vin}'.",
                matched_row=row,
            )
        return result(
            status="Found - Exact Stock Match",
            method="Stock Exact",
            matched_row=row,
        )

    core = stock_numeric_core(dm_stock)
    stripped_candidates = stock_numeric_idx.get(core, []) if core else []
    if len(stripped_candidates) > 1:
        return result(
            status="Needs Review - Multiple Possible Matches",
            method="Stock Stripped",
            notes=f"Multiple Excel rows matched stripped stock core '{core}'.",
        )
    if len(stripped_candidates) == 1:
        row = stripped_candidates[0]
        excel_vin = normalize_vin(row.get(db.vin_header, ""))
        if dm_vin_valid and excel_vin and dm_vin_norm != excel_vin:
            return result(
                status="Needs Review - Stock Match But VIN Conflict",
                method="Stock Stripped",
                notes=f"DeskManager VIN '{dm_vin_norm}' conflicts with Excel VIN '{excel_vin}'.",
                matched_row=row,
            )
        return result(
            status="Found - Stripped Stock Match",
            method="Stock Stripped",
            matched_row=row,
        )

    return result(
        status="Not Found",
        method="None",
        notes="No VIN or stock match found in Excel database.",
    )


def write_output_excel(
    output_path: Path,
    output_rows: List[Dict[str, str]],
    excel_headers: List[str],
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Matched Inventory"

    fixed_headers = [
        "DeskManager Stock Number",
        "DeskManager VIN",
        "Excel Stock Number",
        "Excel VIN",
        "Match Method",
        "Match Status",
        "Review Notes",
    ]
    final_headers = fixed_headers + excel_headers

    ws.append(final_headers)
    for row in output_rows:
        ws.append([row.get(h, "") for h in final_headers])

    wb.save(output_path)


def write_error_log(path: Path, error_lines: List[str]):
    with path.open("w", encoding="utf-8") as f:
        f.write(f"DeskManager Match Errors - {datetime.now().isoformat()}\n")
        f.write("=" * 80 + "\n")
        if not error_lines:
            f.write("No unit-level errors.\n")
            return
        for line in error_lines:
            f.write(line.rstrip() + "\n")


def main():
    print("\n" + "=" * 72)
    print("DESKMANAGER INVENTORY -> EXCEL DATABASE MATCH EXPORT")
    print("=" * 72)

    try:
        db_path = find_database_file()
    except Exception as e:
        print(f"ERROR locating Excel database: {e}")
        return

    print(f"Excel database: {db_path}")
    print(f"Output file: {Path(OUTPUT_XLSX).resolve()}")
    print(f"Error log: {Path(ERROR_LOG_PATH).resolve()}")

    try:
        db = load_excel_database(db_path)
    except Exception as e:
        print(f"ERROR loading Excel database: {e}")
        return

    print(f"Loaded Excel rows: {len(db.rows)}")
    print(f"Detected Excel stock column: {db.stock_header}")
    print(f"Detected Excel VIN column: {db.vin_header}")

    vin_idx, stock_exact_idx, stock_numeric_idx = build_indexes(db)

    output_rows: List[Dict[str, str]] = []
    error_lines: List[str] = []

    summary = {
        "processed": 0,
        "found_vin": 0,
        "found_exact_stock": 0,
        "found_stripped_stock": 0,
        "not_found": 0,
        "needs_review": 0,
        "errors": 0,
    }

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        page = browser.new_page()

        try:
            login(page)
            navigate_to_inventory(page)
            vehicle_hrefs = get_vehicle_hrefs_from_inventory(page)

            if MAX_UNITS > 0:
                vehicle_hrefs = vehicle_hrefs[:MAX_UNITS]
                print(f"Limiting to first {len(vehicle_hrefs)} units due to DESKMANAGER_MAX_UNITS")

            print(f"\nFound {len(vehicle_hrefs)} DeskManager units to process\n")

            for idx, href in enumerate(vehicle_hrefs, 1):
                dm_stock = ""
                dm_vin = ""
                try:
                    page.goto(href, wait_until="domcontentloaded")
                    page.wait_for_timeout(1500)

                    unit = get_stock_vin_from_vehicle(page)
                    dm_stock = unit.get("stock", "")
                    dm_vin = unit.get("vin", "")

                    print(f"[{idx}/{len(vehicle_hrefs)}] Stock='{dm_stock or '-'}' VIN='{dm_vin or '-'}'")

                    match = evaluate_match(
                        dm_stock=dm_stock,
                        dm_vin=dm_vin,
                        db=db,
                        vin_idx=vin_idx,
                        stock_exact_idx=stock_exact_idx,
                        stock_numeric_idx=stock_numeric_idx,
                    )

                    row_out: Dict[str, str] = {
                        "DeskManager Stock Number": dm_stock,
                        "DeskManager VIN": dm_vin,
                        "Excel Stock Number": "",
                        "Excel VIN": "",
                        "Match Method": match["method"],
                        "Match Status": match["status"],
                        "Review Notes": match["notes"],
                    }

                    matched_row = match.get("row")
                    if matched_row:
                        row_out["Excel Stock Number"] = matched_row.get(db.stock_header, "")
                        row_out["Excel VIN"] = matched_row.get(db.vin_header, "")
                        for h in db.headers:
                            row_out[h] = matched_row.get(h, "")
                    else:
                        for h in db.headers:
                            row_out[h] = ""

                    output_rows.append(row_out)

                    status = match["status"]
                    if status.startswith("Found - VIN"):
                        summary["found_vin"] += 1
                    elif status == "Found - Exact Stock Match":
                        summary["found_exact_stock"] += 1
                    elif status == "Found - Stripped Stock Match":
                        summary["found_stripped_stock"] += 1
                    elif status.startswith("Needs Review"):
                        summary["needs_review"] += 1
                    elif status == "Not Found":
                        summary["not_found"] += 1

                    summary["processed"] += 1

                except Exception as e:
                    summary["errors"] += 1
                    summary["processed"] += 1
                    err = f"[{idx}] href={href} stock={dm_stock!r} vin={dm_vin!r} error={e}"
                    error_lines.append(err)
                    error_lines.append(traceback.format_exc())

                    output_rows.append(
                        {
                            "DeskManager Stock Number": dm_stock,
                            "DeskManager VIN": dm_vin,
                            "Excel Stock Number": "",
                            "Excel VIN": "",
                            "Match Method": "Error",
                            "Match Status": "Error",
                            "Review Notes": str(e),
                            **{h: "" for h in db.headers},
                        }
                    )
                    print(f"  ! Error on unit {idx}: {e}")

            write_output_excel(Path(OUTPUT_XLSX), output_rows, db.headers)
            write_error_log(Path(ERROR_LOG_PATH), error_lines)

            print("\n" + "=" * 72)
            print("SUMMARY")
            print("=" * 72)
            print(f"Total DeskManager units processed: {summary['processed']}")
            print(f"Total found by VIN: {summary['found_vin']}")
            print(f"Total found by exact stock number: {summary['found_exact_stock']}")
            print(f"Total found by stripped stock number: {summary['found_stripped_stock']}")
            print(f"Total not found: {summary['not_found']}")
            print(f"Total needs review: {summary['needs_review']}")
            print(f"Total errors: {summary['errors']}")
            print("=" * 72)
            print(f"Output written: {Path(OUTPUT_XLSX).resolve()}")
            print(f"Error log written: {Path(ERROR_LOG_PATH).resolve()}")

        finally:
            browser.close()


if __name__ == "__main__":
    main()
